import asyncio
import csv
import json
import os
from datetime import datetime
from playwright.async_api import async_playwright
import subprocess
import sys

TARGET_URL = "https://nepalstock.com/floor-sheet"
API_URL = "https://nepalstock.com/api/nots/nepse-data/floorsheet"
OUTPUT_DIR = "data"

MAX_RETRIES = 3
PAGE_LOAD_TIMEOUT = 60000

# Column mapping: API field -> display header (businessDate excluded)
# Symbol column comes blank from the API — no fix needed here
COLUMN_MAP = {
    "contractId":       "Transact. No.",
    "stockSymbol":           "Symbol",
    "buyerMemberId":    "Buyer",
    "sellerMemberId":   "Seller",
    "contractQuantity": "Quantity",
    "contractRate":     "Rate",
    "contractAmount":   "Amount",
}

API_FIELDS = list(COLUMN_MAP.keys())
HEADERS    = list(COLUMN_MAP.values())


def extract_trades(data):
    if isinstance(data, list):
        return data
    if isinstance(data, dict):
        if "content" in data and isinstance(data["content"], list):
            return data["content"]
        for key in data.values():
            result = extract_trades(key)
            if result:
                return result
    return None


def extract_market_date(trades):
    sample = trades[0]
    for key in ["tradeDate", "businessDate", "timestamp"]:
        if key in sample:
            raw = sample[key]
            try:
                dt = datetime.fromisoformat(raw.replace("Z", ""))
                return dt.strftime("%m-%d-%Y")
            except:
                pass
    return None


def filter_columns(rows):
    filtered = []
    for row in rows:
        filtered.append({field: row.get(field, "") for field in API_FIELDS})
    return filtered


def save_xlsx(filtered_rows, filepath):
    try:
        import openpyxl
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
        import openpyxl

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Floor Sheet"

    # Header row styling
    header_fill = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)

    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows
    data_font = Font(name="Arial", size=10)

    for row_idx, row in enumerate(filtered_rows, start=2):
        for col_idx, field in enumerate(API_FIELDS, start=1):
            value = row[field]

            # Keep contractId (Transact. No.) as text to prevent number truncation
            if field == "contractId":
                cell = ws.cell(row=row_idx, column=col_idx, value=str(value))
                cell.number_format = "@"
            else:
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

            cell.font = data_font

    # Auto-fit column widths
    col_widths = {"Transact. No.": 16, "Symbol": 10, "Buyer": 8,
                  "Seller": 8, "Quantity": 12, "Rate": 12, "Amount": 16}
    for col_idx, header in enumerate(HEADERS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = \
            col_widths.get(header, 14)

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(filepath)


async def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(
            ignore_https_errors=True,
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120 Safari/537.36"
        )
        page = await context.new_page()

        session = {"auth": None, "id": None}

        def sniff(request):
            if "floorsheet" in request.url and request.method == "POST":
                auth = request.headers.get("authorization")
                post_data = request.post_data
                if auth and post_data and not session["auth"]:
                    try:
                        payload = json.loads(post_data)
                        session["auth"] = auth
                        session["id"] = payload.get("id")
                        print("[✓] Token captured")
                    except:
                        pass

        page.on("request", sniff)

        print(f"Opening {TARGET_URL}")
        for attempt in range(MAX_RETRIES):
            try:
                await page.goto(TARGET_URL, wait_until="domcontentloaded", timeout=PAGE_LOAD_TIMEOUT)
                break
            except Exception as e:
                print(f"Retry {attempt+1} failed: {e}")
                if attempt < MAX_RETRIES - 1:
                    await asyncio.sleep(5)
        else:
            print("❌ Failed to load page after retries")
            await browser.close()
            return

        await asyncio.sleep(3)

        if not session["auth"]:
            print("❌ Token capture failed")
            await browser.close()
            return

        all_trades = []
        page_num = 0
        page_size = 500

        while True:
            print(f"Fetching page {page_num}...")

            js_fetch = f"""
            async () => {{
                const res = await fetch("{API_URL}?page={page_num}&size={page_size}&sort=contractId,desc", {{
                    method: "POST",
                    headers: {{
                        "accept": "application/json, text/plain, */*",
                        "authorization": "{session['auth']}",
                        "content-type": "application/json",
                        "referer": "{TARGET_URL}"
                    }},
                    body: JSON.stringify({{ "id": {session['id']} }}),
                    credentials: "include"
                }});

                let body = null;
                try {{ body = await res.json(); }} catch (e) {{}}

                return {{ status: res.status, body }};
            }}
            """

            result = await page.evaluate(js_fetch)

            if result["status"] == 401:
                print("🔄 Token expired. Reloading...")
                session["auth"] = None
                session["id"] = None
                await page.reload(wait_until="domcontentloaded")
                await asyncio.sleep(3)
                continue

            if result["status"] != 200:
                break

            data = result["body"]
            trades = extract_trades(data)

            if not trades:
                break

            all_trades.extend(trades)
            print(f"  → {len(trades)} rows (Total: {len(all_trades)})")

            page_num += 1
            await asyncio.sleep(1)

        await browser.close()

        if not all_trades:
            print("❌ No trades found - Market was closed (holiday/weekend)")
            return

        print(f"✅ Found {len(all_trades)} trades")

        unique = {t.get("contractId"): t for t in all_trades if t.get("contractId")}
        rows = list(unique.values())

        filtered_rows = filter_columns(rows)

        market_date = extract_market_date(rows)
        if not market_date:
            print("❌ Could not extract market date")
            return

        print(f"Market date: {market_date}")

        filename = f"{market_date}.xlsx"
        filepath = os.path.join(OUTPUT_DIR, filename)

        if os.path.exists(filepath):
            print(f"⚠️ File already exists: {filepath}")
            return

        save_xlsx(filtered_rows, filepath)
        print(f"✅ Saved {len(filtered_rows)} rows to {filepath}")


if __name__ == "__main__":
    asyncio.run(main())
